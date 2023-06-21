from openpyxl import Workbook, load_workbook

class GenericParser:
    input_path : str
    workbook : Workbook
    def __init__(self) -> None:
        load_workbook("C:\Projekte\E2KAO\source\knx.xlsx")
        pass

    def loadWorkbook(self) -> None:
        self.workbook = load_workbook(filename="source/knx.xlsx", data_only=True)

    def loadSheet(self) -> None:
        print("should not be called")