from ast import List
import datetime
from openpyxl import Workbook

from openhab.configParser.genericParser import GenericParser
from openhab.configParser.mapping.modusbusItemMapping import ModbusItemMapping
from openhab.configParser.mapping.modusbusPollerMapping import ModbusPollerMapping
from openhab.items.item_ import Item
from openhab.supportClasses.fontAweSomeIcon import FontAweSomeIcon
from openhab.supportClasses.openhabBindingType import OpenhabBindingType
from openhab.supportClasses.openhabGroup import OpenhabGroup
from openhab.supportClasses.openhabPhysicalUnit import OpenhabPhysicalUnit
from openhab.supportClasses.openhabOutputFormat import OpenhabOutputFormat
from openhab.supportClasses.openhabBinding import OpenhabBinding
from openhab.supportClasses.openhabItemType import OpenhabItemType
from openhab.supportClasses.openhabTag import OpenhabTag
from openhab.things.modbusPoller import ModbusPoller
from openhab.things.modbusThing import ModbusThing


class ModbusParser(GenericParser):

    def __init__(self) -> None:
        self.sheet = None
        self.items: List[Item] = []
        self.things: List[ModbusThing] = []
        self.poller: List[ModbusPoller] = []
        super().__init__()

    def loadSheet(self) -> None:
        # @todo magix number
        self.sheet = self.workbook["MODBUS"]

    def readItems(self) -> None:
        print(str(datetime.datetime.now()) + " - Start parse Modbus Items.")

        for row in self.sheet.iter_rows(min_row=2, values_only=True):

            group = OpenhabGroup()
            group.addGroup(row[ModbusItemMapping.GROUP1])
            group.addGroup(row[ModbusItemMapping.GROUP2])
            group.addGroup(row[ModbusItemMapping.GROUP3])
            group.addGroup(row[ModbusItemMapping.GROUP4])

            self.items.append(Item(OpenhabItemType(row[ModbusItemMapping.TYPE]),
                                   "_".join(filter(bool, [
                                       row[ModbusItemMapping.LOCATION],
                                       row[ModbusItemMapping.ACCESS],
                                       row[ModbusItemMapping.NAME],
                                       row[ModbusItemMapping.EXTENTION],
                                       row[ModbusItemMapping.FUNCTION]])).replace(" ", "_"),
                                   row[ModbusItemMapping.LABEL],
                                   FontAweSomeIcon(
                                       row[ModbusItemMapping.ICON]),
                                   group,
                                   OpenhabTag(row[ModbusItemMapping.TAG]),
                                   OpenhabPhysicalUnit(
                                       row[ModbusItemMapping.EINHEIT]),
                                   OpenhabOutputFormat( row[ModbusItemMapping.FORMAT]),
                                   OpenhabBinding(OpenhabBindingType.MODBUS),
                                   ModbusItemMapping.PERSISTENCE
                                   ))

        print(str(datetime.datetime.now()) + " - End parse Modbus Items.")

    def readThings(self) -> None:
        print(str(datetime.datetime.now()) + " - Start parse Modbus Things.")

        for row in self.sheet.iter_rows(min_row=2, values_only=True):

            self.things.append(ModbusThing("_".join(filter(bool, [
                row[ModbusItemMapping.LOCATION],
                row[ModbusItemMapping.ACCESS],
                row[ModbusItemMapping.NAME],
                row[ModbusItemMapping.EXTENTION],
                row[ModbusItemMapping.FUNCTION]])).replace(" ", "_"),
                row[ModbusItemMapping.LABEL],
                row[ModbusItemMapping.ADDRESS],
                row[ModbusItemMapping.READVALUETYPE],
                row[ModbusItemMapping.POLLER]
            ))

        print(str(datetime.datetime.now()) + " - End parse Modbus Poller.")
        self.readPoller()

    def readPoller(self) -> None:
        print(str(datetime.datetime.now()) + " - Start parse Modbus Poller.")

        self.sheet = self.workbook["MODBUS_POLLER"]

        for row in self.sheet.iter_rows(min_row=2, values_only=True):

            self.poller.append(ModbusPoller("_".join(filter(bool, [
                row[ModbusPollerMapping.LOCATION],
                row[ModbusPollerMapping.ACCESS],
                row[ModbusPollerMapping.NAME],
                row[ModbusPollerMapping.EXTENTION],
                row[ModbusPollerMapping.FUNCTION]])).replace(" ", "_"),
                row[ModbusPollerMapping.LABEL],
                row[ModbusPollerMapping.ADDRESS],
                row[ModbusPollerMapping.LENGTH],
                row[ModbusPollerMapping.TYPE],
                row[ModbusPollerMapping.ID]
            ))

        print(str(datetime.datetime.now()) + " - End parse Modbus Poller.")
