import csv

from openpyxl.reader.excel import load_workbook

from openhab.openhab_generic import OpenhabGeneric

header = ["Group name","Main","Middle","Sub","Central","Unfiltered","Description","DatapointType","Security"]

class KnxCsvWritter:
    def __init__(self) -> None:
        self.pid = open('knx.csv', 'w',encoding="utf-8",newline='')

        self.spamwriter = csv.writer(self.pid, delimiter=',',
                            quotechar='\"', quoting=csv.QUOTE_ALL)
        self.spamwriter.writerow(header)
        self.read_maingroup()
    	
    def write(self,name, main, middle, sub):
        self.spamwriter.writerow([OpenhabGeneric.__remove_umlaut__(name),main,middle,sub,"","","","","Auto"])

    def close(self):
        self.pid.close()

    def read_maingroup(self):
        workbook = load_workbook(filename="source/knx.xlsx")
        sheet = workbook["Hauptgruppen_Bezeichner"]
        i = 0
        for row in sheet.iter_rows(min_row=1, values_only=True):
            self.spamwriter.writerow([OpenhabGeneric.__remove_umlaut__(row[0]),i,"","","","","","","Auto"])
            self.read_middlegroup(i)
            i+=1
        workbook.close()
        
    def read_middlegroup(self, index):
        workbook = load_workbook(filename="source/knx.xlsx")
        sheet = workbook["Mittelgruppen_Bezeichner"]
        i = 0
        for row in sheet.iter_rows(min_row=1, values_only=True):
            self.spamwriter.writerow([OpenhabGeneric.__remove_umlaut__(row[0]),index,i,"","","","","","Auto"])
            i+=1

        workbook.close()