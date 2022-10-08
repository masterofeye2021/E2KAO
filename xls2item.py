from cProfile import label
import json
from operator import iconcat
import time
from openpyxl import Workbook, load_workbook
from knx.group_address import GroupAddress
import item_mapping
from knx.knx_csv import KnxCsvWritter
from oiw import OpenhabItem, OpenhabItemWritter
from openhab.channel.knx_channel import KnxChannel
from openhab.channel.modbus_channel import ModbusChannel
from openhab.equipment.equipment import Equipment
from openhab.items.ekey_item import EkeyItem
from openhab.items.knx_item import KnxItem
from openhab.items.modbus_item import ModbusItem
from openhab.items.network_item import NetworkItem

from openhab.openhab_generic import OpenhabGeneric
from openhab.persistence import Persistence
import api_caller
from openhab.thing_map import ThingMap


sheet_name_point : str = "KNX"
sheet_name_equipment : str = "Equipment"

def xls2item(path :str):

   #Einlesen und verarbeiten der Items und Channels Definitionen
   items = handle_items()

   #Einlesen und verarbeiten der Equipment Definitionen
   equipments = handle_equipment()

   #KNX ETS CSV Format erzeugen
   kcw = KnxCsvWritter()
   #kcw.write(oi.name.get_name("ga"), group_address[0].main, group_address[0].middle, group_address[0].sub)
   kcw.close()
   
   #Openhab Item File erzeugen
   writter = OpenhabItemWritter()
   writter.write_items(equipments, items)

   per = Persistence()
   per.write_persisence(items)


      
   

def handle_items() -> dict:
      start = time.time()
      knxItems = []
      modbusItems = []
      ekeyItems = []
      networkItems = []
      
      thingMap = ThingMap()
      itemMap = dict()

      workbook = load_workbook(filename="source/knx.xlsx", data_only=True)
      sheet = workbook[sheet_name_point]

      for row in sheet.iter_rows(min_row=2, values_only=True):
         i = []
         if row[item_mapping.BINDING] == "KNX":
            i = KnxItem(row)
            c = KnxChannel(row, i, thingMap)
            i.set_bound_to(c.uuidChannel,row[item_mapping.TRANSFORM])
            knxItems.append(i)    
         elif row[item_mapping.BINDING] == "MODBUS":
            i = ModbusItem(row)
            c = ModbusChannel(row)
            i.set_bound_to(c.uuidChannel,row[item_mapping.TRANSFORM])
            modbusItems.append(i)
         elif row[item_mapping.BINDING] == "EKEY":
            ekeyItems.append(EkeyItem(row))
         elif row[item_mapping.BINDING] == "NETWORK":
            networkItems.append(NetworkItem(row))

      workbook.close()

      itemMap["knx"] = knxItems
      itemMap["modbus"] = modbusItems
      itemMap["ekey"] = ekeyItems
      itemMap["network"] = networkItems


      thing = thingMap.get_thing("knx")
      uid = thing["UID"]

      api_caller.delete_channels(uid,OpenhabGeneric.__remove_umlaut__(json.dumps(thing,ensure_ascii=False)))
      api_caller.delete_channels(uid,OpenhabGeneric.__remove_umlaut__(json.dumps(thing,ensure_ascii=False)))
      api_caller.change_thing(uid,OpenhabGeneric.__remove_umlaut__(json.dumps(thing,ensure_ascii=False)))

      ende = time.time()
      print('Handle Items finished in {:5.3f}s'.format(ende-start))
   
      return itemMap
   
def handle_equipment() -> list[Equipment]:
      start = time.time()
      
      euipments = []
      
      workbook = load_workbook(filename="source/knx.xlsx", data_only=True)
      sheet = workbook[sheet_name_equipment]
   
      for row in sheet.iter_rows(min_row=2, values_only=True):
         euipments.append(Equipment(row))
      
      workbook.close()

      ende = time.time()

      print('Handle Equipments finished in {:5.3f}s'.format(ende-start))
      
      return euipments






